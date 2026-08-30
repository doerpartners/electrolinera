"""
Exporta el business case completo de una ubicación evaluada (salida de
scoring.Engine.analyze_point) a un .xlsx con varias hojas: calidad del sitio y racional,
flujo financiero a 9 años, electricidad (tarifa real por ubicación), comisiones, y
contexto/características de la ubicación.

Nota: los valores son un snapshot calculado por la app al momento de exportar — no son
fórmulas vivas de Excel (la fuente de verdad es la app; este archivo es un reporte, no
un modelo para editar y recalcular). Usa openpyxl, ya una dependencia del proyecto para
el ETL — aquí se usa también en tiempo de ejecución, únicamente para este endpoint.
"""
import io
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from . import config

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11)
BOLD = Font(name=FONT_NAME, bold=True)
NORMAL = Font(name=FONT_NAME)
ITALIC_MUTED = Font(name=FONT_NAME, italic=True, color="666666")
WRAP = Alignment(wrap_text=True, vertical="top")
USD = '$#,##0'
USD2 = '$#,##0.0000'
PCT = '0.0%'


def _ws_default_font(ws):
    ws.sheet_view.showGridLines = False


def _kv(ws, row, label, value, number_format=None, label_bold=True):
    ws.cell(row=row, column=1, value=label).font = BOLD if label_bold else NORMAL
    c = ws.cell(row=row, column=2, value=value)
    c.font = NORMAL
    if number_format:
        c.number_format = number_format
    return row + 1


def _header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    return row + 1


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


OPEX_LABELS = {
    "electricidad_energia": "Electricidad (energía, 3 periodos)",
    "electricidad_demanda": "Electricidad (demanda contratada)",
    "electricidad_cargo_fijo": "Electricidad (cargo fijo)",
    "electricidad_dap": "DAP (alumbrado público)",
    "comision_bancaria": "Comisión bancaria / pasarela",
    "mantenimiento": "Mantenimiento",
    "plataforma_software": "Plataforma (software)",
    "servicio": "Servicio (agregado)",
}


def _sheet_resumen(ws, a):
    ws.title = "Resumen"
    _ws_default_font(ws)
    biz = a["business_case"]
    q = a["query"]

    ws.cell(row=1, column=1, value="Business Case — CS Energy MX").font = TITLE_FONT
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.cell(row=2, column=1, value=f"Generado {ts} — snapshot calculado por la app, no un modelo editable.").font = ITALIC_MUTED

    row = 4
    ws.cell(row=row, column=1, value="Ubicación").font = SECTION_FONT
    row += 1
    row = _kv(ws, row, "Latitud / Longitud", f"{q['lat']}, {q['lon']}")
    row = _kv(ws, row, "Metrópoli", q.get("metro") or "fuera de las 6 metrópolis cubiertas")
    row = _kv(ws, row, "Estado", q.get("state") or "n/d")
    row = _kv(ws, row, "Radio de análisis", f"{q['radius_km']} km ({'urbano' if q['density_tier']=='urbano' else 'rural/disperso'}, automático)")

    row += 1
    ws.cell(row=row, column=1, value="Calidad del sitio").font = SECTION_FONT
    row += 1
    row = _kv(ws, row, "Score", f"{a['score']}/100")
    row = _kv(ws, row, "Veredicto", f"{a['verdict']} — {a['verdict_msg']}")
    row += 1
    ws.cell(row=row, column=1, value="Subscore").font = BOLD
    ws.cell(row=row, column=2, value="Valor (0-100)").font = BOLD
    ws.cell(row=row, column=3, value="Peso").font = BOLD
    row += 1
    sub_lbl = {"demand": "Demanda", "gap": "Brecha oferta/demanda", "ses": "Nivel socioeconómico",
               "retail_anchor": "Ancla comercial", "tesla_opportunity": "Oportunidad Tesla"}
    for k, v in a["subscores"].items():
        ws.cell(row=row, column=1, value=sub_lbl.get(k, k))
        ws.cell(row=row, column=2, value=v)
        c = ws.cell(row=row, column=3, value=a["weights"][k])
        c.number_format = PCT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Business case — resumen (1 set de 6 cargadores)").font = SECTION_FONT
    row += 1
    row = _kv(ws, row, "CapEx total", biz["capex_total"], USD)
    row = _kv(ws, row, "Cargadores", biz["chargers"])
    row = _kv(ws, row, "Payback", f"{biz['payback_years']} años ({biz['break_even_months']} meses)" if biz["payback_years"] else "no rentable en 9 años")
    row = _kv(ws, row, "ROI a 9 años", biz["roi"], PCT if isinstance(biz["roi"], float) else None)
    row = _kv(ws, row, f"NPV @ {biz['discount_rate_pct']*100:.0f}%", biz["npv"], USD)
    row = _kv(ws, row, "Valor residual (año 9)", biz["residual_value"], USD)
    row = _kv(ws, row, "Ingreso año 1", biz["revenue_annual"], USD)
    row = _kv(ws, row, "OpEx año 1", biz["opex_annual"], USD)
    row = _kv(ws, row, "Utilidad bruta año 1", biz["gross_profit_annual"], USD)

    row += 1
    ws.cell(row=row, column=1, value="Racional de la decisión").font = SECTION_FONT
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 8, end_column=6)
    cell = ws.cell(row=row, column=1, value=a["rationale"])
    cell.alignment = WRAP
    cell.font = NORMAL
    row += 10
    ws.cell(row=row, column=1, value=a["disclaimer"]).font = ITALIC_MUTED

    _autosize(ws, [30, 22, 12, 10, 10, 10])


def _sheet_flujo(ws, a):
    biz = a["business_case"]
    years = biz["years"]
    opex_keys = list(years[0]["opex_breakdown"].keys())
    headers = (["Año", "Utilización", "Horas activo/cargador/día", "Ingreso"]
               + [OPEX_LABELS.get(k, k) for k in opex_keys]
               + ["OpEx total", "Utilidad bruta", "Utilidad local (16%)", "Utilidad inversionista",
                  "Valor residual", "Utilidad acumulada inversionista"])
    row = _header_row(ws, 1, headers)
    for y in years:
        col = 1
        ws.cell(row=row, column=col, value=y["year"]); col += 1
        c = ws.cell(row=row, column=col, value=y["utilization"]); c.number_format = PCT; col += 1
        ws.cell(row=row, column=col, value=round(y["utilization"] * 24, 1)); col += 1
        ws.cell(row=row, column=col, value=y["revenue"]).number_format = USD; col += 1
        for k in opex_keys:
            ws.cell(row=row, column=col, value=y["opex_breakdown"].get(k, 0)).number_format = USD
            col += 1
        ws.cell(row=row, column=col, value=y["opex"]).number_format = USD; col += 1
        ws.cell(row=row, column=col, value=y["gross_profit"]).number_format = USD; col += 1
        ws.cell(row=row, column=col, value=y["landlord_profit"]).number_format = USD; col += 1
        ws.cell(row=row, column=col, value=y["investor_profit"]).number_format = USD; col += 1
        ws.cell(row=row, column=col, value=y["residual_value"]).number_format = USD; col += 1
        ws.cell(row=row, column=col, value=y["cumulative_investor_profit"]).number_format = USD; col += 1
        row += 1
    ws.cell(row=row + 1, column=1,
            value="Horas activo/cargador/día = utilización × 24h (reparto uniforme entre los 6 cargadores).").font = ITALIC_MUTED
    _autosize(ws, [6, 11, 14, 13] + [16] * len(opex_keys) + [13, 14, 14, 16, 13, 18])


def _sheet_electricidad(ws, a):
    lf = a["business_case"]["local_factors"]
    ph = config.ELECTRICITY["period_hours"]
    row = 1
    ws.cell(row=row, column=1, value="Tarifa eléctrica real por ubicación (CFE, GDMTH)").font = SECTION_FONT
    row += 2
    row = _kv(ws, row, "División CFE", lf["electricity_division"])
    row = _kv(ws, row, "Confianza", lf["electricity_confidence"])
    row = _kv(ws, row, "Fuente", lf["electricity_source"] or "n/d")
    row = _kv(ws, row, "Vigencia (as of)", lf["electricity_as_of"] or "n/d")
    row += 1
    ws.cell(row=row, column=1, value="Periodo").font = BOLD
    ws.cell(row=row, column=2, value="Horas/día").font = BOLD
    ws.cell(row=row, column=3, value="USD/kWh").font = BOLD
    row += 1
    for label, hkey, rkey in [("Punta", "punta", "electricity_punta_usd_kwh"),
                              ("Intermedia", "intermedia", "electricity_intermedia_usd_kwh"),
                              ("Base", "base", "electricity_base_usd_kwh")]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=ph[hkey])
        ws.cell(row=row, column=3, value=lf[rkey]).number_format = USD2
        row += 1
    row += 1
    row = _kv(ws, row, "Cargo por demanda (capacidad contratada)", f"{lf['electricity_demand_usd_kw_month']} USD/kW/mes")
    row = _kv(ws, row, "Precio al usuario", lf["price_per_kwh_user"], USD2)
    row = _kv(ws, row, "Margen de contribución", a["business_case"]["contribution_margin_per_kwh"], USD2)
    row = _kv(ws, row, "Costo de servicio", a["business_case"]["service_cost_per_kwh"], USD2)
    row = _kv(ws, row, "Inflación anual aplicada", lf["inflation_pct"], PCT)
    row = _kv(ws, row, "Crecimiento parque EV/PHEV usado en la proyección", lf["ev_fleet_growth_pct_yoy"], PCT)
    row = _kv(ws, row, "Utilización efectiva (score del sitio)", lf["utilization"], PCT)
    _autosize(ws, [42, 16, 14])


def _sheet_comisiones(ws, a):
    com = a["business_case"]["commissions"]
    row = 1
    ws.cell(row=row, column=1, value="Esquema de comisiones (pago único, informativo)").font = SECTION_FONT
    row += 1
    ws.cell(row=row, column=1, value="No se restan del ROI/NPV del inversionista mostrados en Resumen.").font = ITALIC_MUTED
    row += 2
    row = _kv(ws, row, "VIP (5% del CapEx)", com["vip_usd"], USD)
    row = _kv(ws, row, "Vendedor", com["vendedor_usd"], USD)
    row = _kv(ws, row, "Arquitecto", com["arquitecto_usd"], USD)
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row + 3, end_column=4)
    ws.cell(row=row, column=1, value=com["recurring_note"]).alignment = WRAP
    _autosize(ws, [28, 14, 10, 10])


def _sheet_ubicacion(ws, a):
    ch = a["chargers"]
    veh = a["estimation"]
    nse = a["nse"]
    row = 1
    ws.cell(row=row, column=1, value="Ubicación y características").font = SECTION_FONT
    row += 2

    ws.cell(row=row, column=1, value="Cargadores en el radio").font = BOLD
    row += 1
    for label, key in [("Total", "total"), ("Públicos", "public"), ("Carga rápida", "fast"),
                       ("Tesla", "tesla"), ("Retail/anclaje", "retail"),
                       ("Observados en campo (residencial)", "residential_observed")]:
        row = _kv(ws, row, label, ch[key], label_bold=False)
    row = _kv(ws, row, "EV por cargador público", ch["ev_per_public_charger"] if ch["ev_per_public_charger"] is not None else "n/d", label_bold=False)

    row += 1
    ws.cell(row=row, column=1, value="Parque vehicular estimado (modelado)").font = BOLD
    row += 1
    row = _kv(ws, row, "Autos estimados", veh["cars_est"], label_bold=False)
    row = _kv(ws, row, "Eléctricos/PHEV estimados", veh["ev_est"], label_bold=False)
    row = _kv(ws, row, "% penetración EV local", veh["ev_penetration_pct"] / 100, PCT, label_bold=False)
    row = _kv(ws, row, "Cargadores en casa estimados", veh["home_chargers_est"], label_bold=False)
    row = _kv(ws, row, "Propensión a carga en casa (NSE)", veh["home_charging_propensity"], PCT, label_bold=False)
    row = _kv(ws, row, "Zona urbana (según metros cubiertos)", "Sí" if veh["urban"] else "No", label_bold=False)
    asm = veh["assumptions"]
    row = _kv(ws, row, "Autos/km² asumidos", asm["cars_per_km2"], label_bold=False)
    row = _kv(ws, row, "Multiplicador adopción estatal", asm["state_adoption_mult"], label_bold=False)
    row = _kv(ws, row, "Multiplicador NSE", asm["ses_mult"], label_bold=False)

    row += 1
    ws.cell(row=row, column=1, value="Nivel socioeconómico").font = BOLD
    row += 1
    row = _kv(ws, row, "Índice NSE (0-1)", nse["index"], label_bold=False)
    row = _kv(ws, row, "Fuente", "Polígono oficial" if nse["source"] == "polygon" else "Proxy por señales locales", label_bold=False)
    row = _kv(ws, row, "Zona / NSE", f"{nse['zone'] or 'n/d'} / {nse['nse'] or 'n/d'}", label_bold=False)

    if a["connector_types"]:
        row += 1
        ws.cell(row=row, column=1, value="Tipos de conector cercanos").font = BOLD
        row += 1
        for k, v in a["connector_types"]:
            row = _kv(ws, row, k, v, label_bold=False)

    if a["vehicles_seen"]:
        row += 1
        ws.cell(row=row, column=1, value="Vehículos vistos en la zona (reviews)").font = BOLD
        row += 1
        for k, v in a["vehicles_seen"]:
            row = _kv(ws, row, k, v, label_bold=False)

    fs = a["field_observations"]
    if fs["count"]:
        row += 1
        ws.cell(row=row, column=1, value="Levantamiento de campo").font = BOLD
        row += 1
        row = _kv(ws, row, "Sitios", ", ".join(fs["sites"]), label_bold=False)
        row = _kv(ws, row, "Cargadores operativos", fs["units_operational"], label_bold=False)
        row = _kv(ws, row, "Fuera de servicio", fs["units_out_of_service"], label_bold=False)
        row = _kv(ws, row, "Planeados", fs["units_planned"], label_bold=False)
        row = _kv(ws, row, "Cajones de estacionamiento", fs["parking_spaces"] or "n/d", label_bold=False)
        row = _kv(ws, row, "EV observados (dato real)", fs["ev_observed"] or "n/d", label_bold=False)

    if a["insights"]:
        row += 1
        ws.cell(row=row, column=1, value="Insights").font = BOLD
        row += 1
        for txt in a["insights"]:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            ws.cell(row=row, column=1, value=txt).alignment = WRAP
            row += 1

    if a["nearest_chargers"]:
        row += 1
        ws.cell(row=row, column=1, value="Cargadores más cercanos").font = BOLD
        row += 1
        row = _header_row(ws, row, ["Nombre", "Clase", "Tesla", "Rápido", "Dist. (km)", "Conectores"])
        for p in a["nearest_chargers"]:
            ws.cell(row=row, column=1, value=p["name"])
            ws.cell(row=row, column=2, value=p["class"])
            ws.cell(row=row, column=3, value="Sí" if p["tesla"] else "No")
            ws.cell(row=row, column=4, value="Sí" if p["fast"] else "No")
            ws.cell(row=row, column=5, value=p["dist_km"])
            ws.cell(row=row, column=6, value=", ".join(p["connectors"]))
            row += 1

    _autosize(ws, [34, 40, 10, 10, 12, 30])


def build_business_case_xlsx(a):
    """a: dict devuelto por scoring.Engine.analyze_point(). Regresa bytes .xlsx."""
    wb = openpyxl.Workbook()
    _sheet_resumen(wb.active, a)
    _sheet_flujo(wb.create_sheet("Flujo 9 años"), a)
    _sheet_electricidad(wb.create_sheet("Electricidad"), a)
    _sheet_comisiones(wb.create_sheet("Comisiones"), a)
    _sheet_ubicacion(wb.create_sheet("Ubicación y contexto"), a)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
