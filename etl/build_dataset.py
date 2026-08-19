#!/usr/bin/env python3
"""
ETL: normaliza los datos crudos (PlugShare + INEGI RAIAVL) hacia JSON limpio
que consume la API. Idempotente: puede correrse cuando lleguen datos nuevos.

Salidas (en data/processed/):
  - chargers.json       Cargadores únicos, geolocalizados y clasificados.
  - demand_state.json   Adopción EV/PHEV por entidad (RAIAVL) + shares AMIA.
  - insights.json       Constantes derivadas de reportes oficiales (ICCT/AMIA).
  - meta.json           Resumen del build.
"""
import json, os, re, glob
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("Falta openpyxl:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT = os.path.join(ROOT, "data", "processed")
os.makedirs(OUT, exist_ok=True)

# --- Ubicación de los archivos fuente ------------------------------------
# Por defecto se leen desde ~/Downloads. Ajusta si mueves los archivos.
DOWNLOADS = os.path.expanduser("~/Downloads")
PLUGSHARE_GLOB = os.path.join(DOWNLOADS, "dataset_plugshare-scraper_*.xlsx")
# Archivos PlugShare adicionales que no siguen el patrón del glob:
EXTRA_PLUGSHARE = [os.path.join(DOWNLOADS, "mérida.xlsx")]
RAIAVL = os.path.join(DOWNLOADS, "RAIAVL_11.xlsx")

# Conectores considerados "residenciales" (carga en casa / bajo nivel)
RESIDENTIAL_CONNECTORS = {"NEMA 14-50", "NEMA TT-30", "Wall", "NEMA 5-15", "NEMA 5-20"}
FAST_CONNECTORS = {"CCS1", "CCS2", "CHAdeMO", "GB/T (Fast)", "NACS (Tesla)"}
RETAIL_AMENITIES = {"Shopping", "Dining", "Grocery"}


def load_headers(ws):
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    return {h: i for i, h in enumerate(header) if h}


def cell(row, idx, key):
    i = idx.get(key)
    if i is None or i >= len(row):
        return None
    v = row[i]
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def collect_multi(row, idx, prefix, maxn=12):
    """Recolecta connectorTypes/0..n, amenities/0..n, etc."""
    out = []
    for k in range(maxn):
        v = cell(row, idx, f"{prefix}/{k}")
        if v:
            out.append(v)
    return out


def to_bool(v):
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().lower() in ("true", "1", "yes", "sí", "si")
    return bool(v)


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def classify_charger(connectors, amenities, is_fast, network, station_manufacturers):
    """Devuelve ('public'|'residential', tesla_bool)."""
    tesla = False
    if network and "tesla" in network.lower():
        tesla = True
    if any(m and "tesla" in str(m).lower() for m in station_manufacturers):
        tesla = True
    if any("NACS" in c for c in connectors):
        tesla = True

    conn_set = set(connectors)
    only_residential = conn_set and conn_set.issubset(RESIDENTIAL_CONNECTORS)
    has_retail = bool(set(amenities) & RETAIL_AMENITIES)

    if only_residential and not is_fast and not has_retail:
        klass = "residential"
    else:
        klass = "public"
    return klass, tesla


def build_chargers():
    files = sorted(glob.glob(PLUGSHARE_GLOB)) + [f for f in EXTRA_PLUGSHARE if os.path.exists(f)]
    if not files:
        raise SystemExit(f"No se encontraron archivos PlugShare en {PLUGSHARE_GLOB}")
    seen = {}
    for f in files:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        ws = wb.active
        idx = load_headers(ws)
        for row in ws.iter_rows(min_row=2, values_only=True):
            cid = cell(row, idx, "id")
            lat = to_float(cell(row, idx, "latitude"))
            lon = to_float(cell(row, idx, "longitude"))
            if cid is None or lat is None or lon is None:
                continue
            if cid in seen:
                continue  # dedupe por id (primera aparición)

            connectors = collect_multi(row, idx, "connectorTypes")
            # también de outlets, por si connectorTypes viene vacío
            if not connectors:
                for s in range(8):
                    for o in range(2):
                        c = cell(row, idx, f"stations/{s}/outlets/{o}/connectorType")
                        if c:
                            connectors.append(c)
            connectors = sorted(set(connectors))

            amenities = collect_multi(row, idx, "amenities")
            manuf = [cell(row, idx, f"stations/{s}/manufacturer") for s in range(8)]
            network = cell(row, idx, "network")
            is_fast = to_bool(cell(row, idx, "isFastCharger"))

            klass, tesla = classify_charger(connectors, amenities, is_fast, network, manuf)

            # vehículos vistos en reviews (tipos de vehículo en la zona)
            vehicles = []
            for rv in range(10):
                veh = cell(row, idx, f"reviews/{rv}/vehicle")
                if veh:
                    vehicles.append(veh)

            seen[cid] = {
                "id": str(cid),
                "name": cell(row, idx, "name") or "",
                "address": cell(row, idx, "address") or "",
                "lat": lat,
                "lon": lon,
                "network": network,
                "class": klass,            # public | residential
                "is_fast": bool(is_fast),
                "tesla": bool(tesla),
                "kilowatts": to_float(cell(row, idx, "kilowatts")),
                "station_count": cell(row, idx, "stationCount") or 1,
                "connectors": connectors,
                "amenities": amenities,
                "retail": bool(set(amenities) & RETAIL_AMENITIES),
                "score": to_float(cell(row, idx, "score")),
                "reviews": cell(row, idx, "totalReviews") or 0,
                "vehicles": vehicles,
                "url": cell(row, idx, "url") or "",
            }
        wb.close()
    return list(seen.values())


# Mapa de nombres de entidad INEGI -> clave corta
STATE_KEY = {
    "Ciudad de México": "CDMX", "México": "MEX", "Nuevo León": "NL", "Jalisco": "JAL",
    "Guanajuato": "GTO", "Puebla": "PUE", "Querétaro": "QRO", "Veracruz de Ignacio de la Llave": "VER",
    "Yucatán": "YUC", "Coahuila de Zaragoza": "COA", "Aguascalientes": "AGS",
    "Baja California": "BC", "Baja California Sur": "BCS", "Campeche": "CAM",
    "Colima": "COL", "Chiapas": "CHP", "Chihuahua": "CHH", "Durango": "DUR",
    "Guerrero": "GRO", "Hidalgo": "HID", "Michoacán de Ocampo": "MIC", "Morelos": "MOR",
    "Nayarit": "NAY", "Oaxaca": "OAX", "Quintana Roo": "ROO", "San Luis Potosí": "SLP",
    "Sinaloa": "SIN", "Sonora": "SON", "Tabasco": "TAB", "Tamaulipas": "TAM",
    "Tlaxcala": "TLA", "Zacatecas": "ZAC",
}


def build_demand():
    wb = openpyxl.load_workbook(RAIAVL, data_only=True)
    ws = wb["Tabulado"]
    # Encontrar fila de encabezado
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    for i, r in enumerate(rows):
        if r and r[0] == "Año":
            header_i = i
            break
    agg = defaultdict(lambda: {"bev": 0, "phev": 0, "months": 0})
    cur_state = None
    for r in rows[header_i + 1:]:
        entidad = r[2]
        bev = r[3]
        phev = r[4]
        if entidad:
            cur_state = entidad.strip()
        if cur_state is None:
            continue
        try:
            b = int(bev) if bev is not None else 0
            p = int(phev) if phev is not None else 0
        except (TypeError, ValueError):
            continue
        agg[cur_state]["bev"] += b
        agg[cur_state]["phev"] += p
        agg[cur_state]["months"] += 1
    wb.close()

    out = {}
    total = sum(v["bev"] + v["phev"] for v in agg.values())
    for state, v in agg.items():
        key = STATE_KEY.get(state, state)
        units = v["bev"] + v["phev"]
        out[key] = {
            "entidad": state,
            "bev": v["bev"],
            "phev": v["phev"],
            "ev_units_period": units,
            "share_national": round(units / total, 4) if total else 0,
        }
    return out, total


def build_insights():
    """Constantes derivadas de reportes oficiales (ICCT Market Spotlight 2025,
    AMIA Ene-Abr 2026). Este archivo es el 'cerebro' tunable del sistema:
    aquí se alimentan insights nuevos (marca->NSE, precios, carga en casa)."""
    return {
        "source_notes": [
            "ICCT Market Spotlight, Marzo 2026 — Evolución del mercado EV ligero en México 2025.",
            "AMIA/INEGI, reporte híbridos+PHEV+EV Ene-Abr 2026.",
        ],
        # EVs son producto premium => correlación con NSE alto.
        "avg_price_mxn": {"BEV": 822000, "PHEV": 845000, "ICE": 554000, "HEV": 864000},
        "market_2025": {
            "ldv_total_units": 1620000,
            "ev_share": 0.071, "bev_share": 0.033, "phev_share": 0.038,
            "ev_growth_yoy": 0.56, "phev_growth_yoy": 0.97, "bev_growth_yoy": 0.26,
        },
        # Marca -> segmento dominante y tier de NSE implícito por precio/posicionamiento.
        # ses_tier: 1=medio, 2=medio-alto, 3=alto/premium  (tunable)
        "brand_ses": {
            "Tesla":  {"segment": "BEV", "ses_tier": 3, "note": "Model 3/Y, premium, 19% de BEV"},
            "Volvo":  {"segment": "BEV", "ses_tier": 3, "note": "EX30, premium europeo"},
            "BMW":    {"segment": "PHEV", "ses_tier": 3, "note": "premium"},
            "Mercedes Benz": {"segment": "BEV", "ses_tier": 3, "note": "premium"},
            "BYD":    {"segment": "PHEV", "ses_tier": 2, "note": "líder EV: 56% BEV, 74% PHEV; Dolphin/Song/Shark"},
            "GM":     {"segment": "BEV", "ses_tier": 2, "note": "Spark EV, producción MX"},
            "Ford":   {"segment": "BEV", "ses_tier": 2},
            "JAC":    {"segment": "BEV", "ses_tier": 1, "note": "entrada, producción MX"},
            "Chevrolet": {"segment": "BEV", "ses_tier": 2},
        },
        # Participación de EV por marca (BEV) — ICCT 2025 (aprox).
        "bev_brand_share_2025": {"BYD": 0.56, "Tesla": 0.19, "JAC": 0.02, "Volvo": 0.01},
        # Propensión a cargar en casa por tier NSE (mayor NSE => más carga en casa,
        # => menor demanda pública "de necesidad", pero mayor demanda "destino/retail").
        # Insight tunable — alimentar con datos reales cuando existan.
        "home_charging_propensity": {"1": 0.25, "2": 0.55, "3": 0.80},
        # Participación acumulada de ventas híbrido+PHEV+EV por entidad (AMIA Ene-Abr 2026)
        "amia_state_share_2026": {
            "CDMX": 0.242, "MEX": 0.132, "NL": 0.105, "JAL": 0.085, "GTO": 0.040,
            "PUE": 0.034, "QRO": 0.030, "VER": 0.029, "YUC": 0.025, "COA": 0.024,
        },
    }


def main():
    print("→ Cargadores…")
    chargers = build_chargers()
    print("→ Demanda estatal (RAIAVL)…")
    demand, total = build_demand()
    print("→ Insights (reportes)…")
    insights = build_insights()

    json.dump(chargers, open(os.path.join(OUT, "chargers.json"), "w"), ensure_ascii=False)
    json.dump(demand, open(os.path.join(OUT, "demand_state.json"), "w"), ensure_ascii=False, indent=2)
    json.dump(insights, open(os.path.join(OUT, "insights.json"), "w"), ensure_ascii=False, indent=2)

    pub = sum(1 for c in chargers if c["class"] == "public")
    res = sum(1 for c in chargers if c["class"] == "residential")
    tesla = sum(1 for c in chargers if c["tesla"])
    fast = sum(1 for c in chargers if c["is_fast"])
    retail = sum(1 for c in chargers if c["retail"])
    meta = {
        "chargers_total": len(chargers), "public": pub, "residential": res,
        "tesla_sites": tesla, "fast": fast, "retail_sites": retail,
        "states_with_demand": len(demand), "ev_units_period_total": total,
    }
    json.dump(meta, open(os.path.join(OUT, "meta.json"), "w"), ensure_ascii=False, indent=2)
    print("\nOK. Resumen:")
    print(json.dumps(meta, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
